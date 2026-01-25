import csv
import io
import os
import sys
import tempfile
from datetime import date, timedelta
from decimal import Decimal

import boto3
from account.models import CustomUser
from django.conf import settings
from django.contrib import admin, messages
from django.core.exceptions import ValidationError
from django.db import IntegrityError, transaction
from django.db.models import Count, Sum
from django.db.models.functions import Round
from django.forms import ModelForm
from django.http import HttpResponse, HttpResponseRedirect
from django.template.loader import render_to_string
from django.urls import path, reverse
from django.utils.html import format_html
from django.utils.translation import gettext_lazy as _
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from billing.models import CreditNote
from shipping.models import ShippingDetails

from .forms import (
    OrderItemInlineForm,
    OrderItemInlineFormSet,
    ProductImageAdminForm,
    ProductImageInlineForm,
)
from .models import (
    Cart,
    CartItem,
    CustomUser,
    Order,
    OrderItem,
    Product,
    ProductCategory,
    ProductImage,
    Wishlist,
    WishlistItem,
)


class OrderAdminForm(ModelForm):
    """Custom form for Order admin."""

    class Meta:
        model = Order
        fields = "__all__"

    def clean(self):
        cleaned_data = super().clean()
        return cleaned_data


class ProductImageInline(admin.TabularInline):
    model = ProductImage
    form = ProductImageInlineForm
    extra = 1
    fields = ["image_preview", "image_file", "image_url", "sort_order", "alt_text"]
    readonly_fields = ["image_preview"]
    ordering = ["sort_order"]

    class Media:
        css = {"all": ("admin/css/product_image_inline.css",)}

    def image_preview(self, obj):
        if obj.image_url:
            # Mark the first image as primary visually with green border
            is_primary = obj.sort_order == 0 or (
                obj.product and obj.product.images.first() == obj
            )
            border_style = (
                "border: 3px solid #4CAF50;"
                if is_primary
                else "border: 1px solid #ddd;"
            )

            return format_html(
                '<img src="{}" style="max-width: 100px; max-height: 100px; object-fit: contain; {}; border-radius: 4px; padding: 2px;" />',
                obj.image_url,
                border_style,
            )
        return format_html(
            '<div style="width: 100px; height: 100px; border: 2px dashed #ccc; border-radius: 4px; '
            'display: flex; align-items: center; justify-content: center; color: #999; font-size: 11px; text-align: center;">'
            "No image<br>yet</div>"
        )

    image_preview.short_description = "Preview"


# @admin.register(ProductImage)
# class ProductImageAdmin(admin.ModelAdmin):
#     form = ProductImageAdminForm
#     list_display = ["product", "image_preview_thumb", "sort_order", "is_primary_display", "created_at"]
#     list_filter = ["product"]
#     search_fields = ["product__name", "alt_text"]
#     ordering = ["product", "sort_order"]
#     autocomplete_fields = ["product"]

#     fieldsets = (
#         ('Product', {
#             'fields': ('product',)
#         }),
#         ('Image', {
#             'fields': ('image_file', 'image_url', 'alt_text'),
#             'description': 'Upload an image file or provide a URL. If you upload a file, it will be automatically uploaded to R2 storage.'
#         }),
#         ('Display Settings', {
#             'fields': ('sort_order',),
#             'description': 'The first image (sort_order = 0) is automatically the primary image.'
#         }),
#     )

#     def image_preview_thumb(self, obj):
#         if obj.image_url:
#             is_primary = obj.is_primary
#             border_color = "#4CAF50" if is_primary else "#ddd"
#             return format_html(
#                 '<img src="{}" style="max-width: 50px; max-height: 50px; object-fit: contain; border: 2px solid {}; border-radius: 4px;" />',
#                 obj.image_url,
#                 border_color
#             )
#         return format_html(
#             '<span style="color: #999; font-size: 11px;">No image</span>'
#         )

#     image_preview_thumb.short_description = "Preview"

#     def is_primary_display(self, obj):
#         """Display if this is the primary image (first by sort_order)."""
#         return obj.is_primary

#     is_primary_display.boolean = True
#     is_primary_display.short_description = "Primary"


@admin.register(Product)
class ProductAdmin(admin.ModelAdmin):
    list_display = ["name", "get_price", "get_vat_display", "get_categories"]
    list_filter = ["categories", "vat"]
    filter_horizontal = ["categories"]
    search_fields = ["name"]
    ordering = ["name"]
    inlines = [ProductImageInline]

    fields = ["name", "description", "base_price", "holiday_fee", "vat", "categories"]

    class Media:
        js = ("admin/js/prevent_double_submit.js",)

    def get_queryset(self, request):
        qs = super().get_queryset(request)
        return qs.prefetch_related("categories", "images").distinct()

    def get_price(self, obj):
        """Display the calculated final price."""
        return f"£{obj.price}"

    get_price.short_description = "Final Price"

    def get_vat_display(self, obj):
        """Display VAT percentage."""
        vat_percent = obj.vat_percentage * Decimal("100")
        return f"{vat_percent:.0f}%"

    get_vat_display.short_description = "VAT"

    def get_readable_categories(self, obj):
        return ", ".join(obj.get_categories)

    get_readable_categories.short_description = "Categories"

    def formfield_for_manytomany(self, db_field, request, **kwargs):
        if db_field.name == "categories":
            # Only categories that are not parents (i.e., that are leaf nodes)
            kwargs["queryset"] = ProductCategory.objects.filter(
                subcategories__isnull=True
            ).order_by("parent__name", "name")
        return super().formfield_for_manytomany(db_field, request, **kwargs)

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)

        # Now categories are saved, safe to modify them
        instance = form.instance
        to_add = set()
        for cat in instance.categories.all():
            parent = cat.parent
            while parent:
                to_add.add(parent)
                parent = parent.parent
        if to_add:
            instance.categories.add(*to_add)


class ParentCategoriesFilter(admin.SimpleListFilter):
    title = _("Parent Category")
    parameter_name = "parent_category"

    def lookups(self, request, model_admin):
        # Only show parents that have subcategories
        parents = ProductCategory.objects.filter(subcategories__isnull=False).distinct()
        return [(parent.id, parent.name) for parent in parents]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(parent__id=self.value())
        return queryset


class HolidayFeeFilter(admin.SimpleListFilter):
    title = _("Have holiday_fee")
    parameter_name = "have_holiday_fee"

    def lookups(self, request, model_admin):
        return [
            ("yes", _("Yes")),
            ("no", _("No")),
        ]

    def queryset(self, request, queryset):
        if self.value() == "yes":
            return queryset.filter(holiday_fee__gt=0)
        elif self.value() == "no":
            return queryset.filter(holiday_fee=0)
        return queryset


@admin.register(ProductCategory)
class ProductCategoryAdmin(admin.ModelAdmin):
    list_display = ["name", "description", "parent"]
    list_filter = [ParentCategoriesFilter]
    search_fields = ["name"]
    ordering = ["parent__name", "name"]

    class Media:
        js = ("admin/js/prevent_double_submit.js",)


class CartItemInline(admin.TabularInline):
    model = CartItem
    extra = 0
    fields = ["product", "quantity", "added_date"]
    readonly_fields = ["added_date"]
    autocomplete_fields = ["product"]


@admin.register(Cart)
class CartAdmin(admin.ModelAdmin):
    list_display = ["user", "total_items", "total_price", "created_at", "updated_at"]
    list_filter = ["created_at", "updated_at"]
    search_fields = ["user__name", "user__email"]
    readonly_fields = ["created_at", "updated_at", "total_items", "total_price"]
    inlines = [CartItemInline]

    class Media:
        js = ("admin/js/prevent_double_submit.js",)

    def total_items(self, obj):
        return obj.total_items

    total_items.short_description = "Total Items"

    def total_price(self, obj):
        return f"£{obj.total_price:.2f}"

    total_price.short_description = "Total Price"


@admin.register(CartItem)
class CartItemAdmin(admin.ModelAdmin):
    list_display = ["cart", "product", "quantity", "get_total_price", "added_date"]
    list_filter = ["added_date"]
    search_fields = ["cart__user__name", "product__name"]
    readonly_fields = ["added_date", "get_total_price"]
    autocomplete_fields = ["cart", "product"]

    class Media:
        js = ("admin/js/prevent_double_submit.js",)

    def get_total_price(self, obj):
        return f"£{obj.get_total_price():.2f}"

    get_total_price.short_description = "Total Price"


class WishlistItemInline(admin.TabularInline):
    model = WishlistItem
    extra = 0
    fields = ["product", "added_date"]
    readonly_fields = ["added_date"]
    autocomplete_fields = ["product"]


@admin.register(Wishlist)
class WishlistAdmin(admin.ModelAdmin):
    list_display = ["user", "total_items", "created_at", "updated_at"]
    list_filter = ["created_at", "updated_at"]
    search_fields = ["user__name", "user__email"]
    readonly_fields = ["created_at", "updated_at", "total_items"]
    inlines = [WishlistItemInline]

    class Media:
        js = ("admin/js/prevent_double_submit.js",)

    def total_items(self, obj):
        return obj.total_items

    total_items.short_description = "Total Items"


@admin.register(WishlistItem)
class WishlistItemAdmin(admin.ModelAdmin):
    list_display = ["wishlist", "product", "added_date"]
    list_filter = ["added_date"]
    search_fields = ["wishlist__user__name", "product__name"]
    readonly_fields = ["added_date"]
    autocomplete_fields = ["wishlist", "product"]

    class Media:
        js = ("admin/js/prevent_double_submit.js",)


class DateFilter(admin.SimpleListFilter):
    title = _("Delivery Date")
    parameter_name = "future_date"

    def lookups(self, request, model_admin):
        return [
            ("past 7 days", _("Past 7 days")),
            ("today", _("Today")),
            ("Next 7 days", _("Next 7 days")),
            ("Next 30 days", _("Next 30 days")),
        ]

    def queryset(self, request, queryset):
        today = date.today()
        if self.value() == "past 7 days":
            return queryset.filter(
                delivery_date__gte=today - timedelta(days=7), delivery_date__lte=today
            )
        elif self.value() == "today":
            return queryset.filter(delivery_date=today)
        elif self.value() == "Next 7 days":
            return queryset.filter(
                delivery_date__gte=today, delivery_date__lte=today + timedelta(days=7)
            )
        elif self.value() == "Next 30 days":
            return queryset.filter(
                delivery_date__gte=today, delivery_date__lte=today + timedelta(days=30)
            )
        return queryset


def get_s3_client():
    return boto3.client(
        "s3",
        aws_access_key_id=settings.AWS_ACCESS_KEY_ID,
        aws_secret_access_key=settings.AWS_SECRET_ACCESS_KEY,
        region_name=getattr(settings, "AWS_S3_REGION_NAME", None),
    )


def upload_invoice_to_s3(file_path, s3_key, max_retries=3):
    """
    Upload invoice to S3 with retry logic and error handling.

    Args:
        file_path: Path to the PDF file to upload
        s3_key: S3 key (path) where the file should be stored
        max_retries: Maximum number of retry attempts

    Returns:
        str: The S3 key if successful

    Raises:
        Exception: If upload fails after all retries
    """
    import os
    import time

    s3_client = get_s3_client()
    bucket = settings.AWS_STORAGE_BUCKET_NAME

    # Validate file exists and is readable
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Invoice file not found: {file_path}")

    if not os.access(file_path, os.R_OK):
        raise PermissionError(f"Cannot read invoice file: {file_path}")

    # Retry logic for network issues
    last_exception = None
    for attempt in range(1, max_retries + 1):
        try:
            s3_client.upload_file(file_path, bucket, s3_key)
            return s3_key
        except Exception as e:
            last_exception = e
            if attempt < max_retries:
                # Exponential backoff: wait 1s, 2s, 4s...
                wait_time = 2 ** (attempt - 1)
                time.sleep(wait_time)
            else:
                raise Exception(
                    f"Failed to upload invoice to S3 after {max_retries} attempts: {str(e)}"
                ) from last_exception

    raise Exception(f"Failed to upload invoice to S3: {str(last_exception)}")


@admin.action(description="Create Accounting Invoice + Upload PDF")
def create_and_upload_invoice(modeladmin, request, queryset):
    """
    Create accounting invoices (billing.Invoice) AND upload PDF invoices for selected orders with
    robust error handling,
    memory management, and optimization.

    Features:
    - Reuses font configuration to avoid repeated font subsetting
    - Processes orders in chunks to prevent memory exhaustion
    - Comprehensive error handling with detailed feedback
    - Database transactions for atomicity
    - Retry logic for S3 uploads
    - Memory management with garbage collection
    - Optimized PDF generation settings
    - Progress tracking and validation
    - Immutable accounting invoice snapshot creation (billing.Invoice + line items)
    """
    import gc
    import logging
    import time

    from django.db import transaction

    # Count orders for progress tracking
    order_count = queryset.count()
    if order_count == 0:
        modeladmin.message_user(request, "No orders selected.", level=messages.WARNING)
        return

    # Estimate processing time (roughly 2-3 seconds per order)
    estimated_time = order_count * 2.5
    logger = logging.getLogger(__name__)
    logger.info(
        f"Starting creation and uploading of {order_count} order(s), "
        f"approximate time {estimated_time:.1f} seconds"
    )

    # Optimize queryset to avoid N+1 queries
    optimized_queryset = (
        queryset.select_related("customer", "customer__profile", "address")
        .prefetch_related("items", "items__product")
        .order_by("id")
    )

    # Optimize chunk size based on order count for better throughput
    # Larger chunks = better throughput, but need to balance memory
    if order_count > 50:
        chunk_size = 20  # Larger chunks for big batches
    elif order_count > 20:
        chunk_size = 15
    else:
        chunk_size = min(10, order_count)  # Smaller chunks for small batches

    success_count = 0
    error_count = 0
    errors = []
    start_time = time.time()

    # Process orders in chunks
    for chunk_idx, i in enumerate(range(0, order_count, chunk_size), 1):
        chunk_orders = list(optimized_queryset[i : i + chunk_size])

        for order in chunk_orders:
            try:
                # Validate order has required data
                if not order.customer:
                    raise ValueError(f"Order {order.id} has no customer")

                # Use database transaction for atomicity
                with transaction.atomic():
                    from billing.models import Invoice

                    invoice = Invoice.create_and_publish_from_order(
                        order=order, request=request
                    )

                    # Backward compat: keep the Order.invoice_link in sync for existing UI/code paths
                    if (
                        invoice.invoice_link
                        and order.invoice_link != invoice.invoice_link
                    ):
                        order.invoice_link = invoice.invoice_link
                        order.save(update_fields=["invoice_link"])

                    success_count += 1
                    
                    # Small delay to prevent overwhelming WeasyPrint and reduce crash risk
                    time.sleep(0.1)

            except Exception as e:
                error_count += 1
                error_msg = f"Order {order.id}: {str(e)}"
                errors.append(error_msg)
                logger.error(error_msg, exc_info=True)
                # Continue processing other orders even if one fails
                continue

        # Clean up chunk data and force garbage collection
        del chunk_orders
        gc.collect()

    # Prepare user feedback message
    elapsed_time = time.time() - start_time

    if success_count == order_count:
        message = (
            f"Successfully created and uploaded {success_count} invoice(s) "
            f"in {elapsed_time:.1f} seconds."
        )
        logger.info(message)
        modeladmin.message_user(request, message, level=messages.SUCCESS)
    elif success_count > 0:
        message = (
            f"Created and uploaded {success_count} invoice(s) successfully, "
            f"but {error_count} failed. "
            f"Time taken: {elapsed_time:.1f} seconds."
        )
        logger.info(message)
        modeladmin.message_user(request, message, level=messages.WARNING)

        # Show detailed errors (limit to first 10 to avoid overwhelming UI)
        error_display = "\n".join(errors[:10])
        if len(errors) > 10:
            error_display += f"\n... and {len(errors) - 10} more errors."
        modeladmin.message_user(
            request,
            f"Errors:\n{error_display}",
            level=messages.ERROR,
        )
    else:
        message = (
            f"Failed to create any invoices. {error_count} error(s) occurred. "
            f"Time taken: {elapsed_time:.1f} seconds."
        )
        logger.error(message)
        modeladmin.message_user(request, message, level=messages.ERROR)

        # Show all errors if all failed
        error_display = "\n".join(errors[:20])
        if len(errors) > 20:
            error_display += f"\n... and {len(errors) - 20} more errors."
        modeladmin.message_user(
            request,
            f"Errors:\n{error_display}",
            level=messages.ERROR,
        )


@admin.action(description="Mark selected orders as Pending")
def mark_orders_pending(modeladmin, request, queryset):
    updated = queryset.update(status="pending")
    modeladmin.message_user(
        request,
        f"{updated} order(s) marked as pending.",
        level=messages.SUCCESS,
    )


@admin.action(description="Mark selected orders as Paid")
def mark_orders_paid(modeladmin, request, queryset):
    from billing.models import Invoice
    from django.utils import timezone

    updated = queryset.update(status="paid")
    
    # Update related invoices to mark them as paid
    invoice_updated = 0
    for order in queryset.select_related("invoice"):
        invoice = getattr(order, "invoice", None)
        if invoice and invoice.status != Invoice.Status.PAID:
            invoice.amount_paid = invoice.total_amount
            invoice.status = Invoice.Status.PAID
            invoice.paid_at = timezone.now()
            invoice.save(update_fields=["amount_paid", "status", "paid_at"])
            invoice_updated += 1
    
    message = f"{updated} order(s) marked as paid."
    if invoice_updated > 0:
        message += f" {invoice_updated} invoice(s) updated."
    
    modeladmin.message_user(
        request,
        message,
        level=messages.SUCCESS,
    )


@admin.action(description="Mark selected orders as Cancelled")
def mark_orders_cancelled(modeladmin, request, queryset):
    updated = queryset.update(status="cancelled")
    modeladmin.message_user(
        request,
        f"{updated} order(s) marked as cancelled.",
        level=messages.SUCCESS,
    )


@admin.action(description="Create Credit Note")
def create_credit_note_and_void_invoice(modeladmin, request, queryset):
    """
    Create credit notes for selected orders' invoices and void the invoices.
    This action:
    - Creates a credit note with all snapshots copied from the invoice
    - Generates and uploads the credit note PDF to S3
    - Voids the original invoice
    - Marks the order as cancelled
    """
    success_count = 0
    error_count = 0
    skipped_count = 0
    errors = []

    for order in queryset.select_related("invoice"):
        try:
            # Check if order has an invoice
            invoice = getattr(order, "invoice", None)
            if not invoice:
                skipped_count += 1
                errors.append(f"Order #{order.id}: No invoice exists")
                continue

            # Check if invoice already has a credit note
            try:
                existing_credit_note = invoice.credit_note
                if existing_credit_note:
                    skipped_count += 1
                    errors.append(
                        f"Order #{order.id}: Invoice already has Credit Note #{existing_credit_note.credit_note_number}"
                    )
                    continue
            except CreditNote.DoesNotExist:
                pass

            # Create credit note and void invoice
            credit_note = CreditNote.create_and_publish_from_invoice(
                invoice=invoice,
                reason="Order cancelled/refunded via admin action",
                request=request,
            )

            # Mark order as cancelled
            order.status = "cancelled"
            order.save(update_fields=["status"])

            success_count += 1

        except Exception as e:
            error_count += 1
            errors.append(f"Order #{order.id}: {str(e)}")

    # Build result message
    if success_count > 0:
        modeladmin.message_user(
            request,
            f"Successfully created {success_count} credit note(s) and voided invoice(s).",
            level=messages.SUCCESS,
        )

    if skipped_count > 0:
        modeladmin.message_user(
            request,
            f"Skipped {skipped_count} order(s): {'; '.join(errors[:5])}{'...' if len(errors) > 5 else ''}",
            level=messages.WARNING,
        )

    if error_count > 0:
        modeladmin.message_user(
            request,
            f"Failed to process {error_count} order(s). Check logs for details.",
            level=messages.ERROR,
        )


@admin.action(description="Get the total income")
def calculate_sum(modeladmin, request, queryset):
    total_sum = sum(order.total_price for order in queryset)
    modeladmin.message_user(
        request,
        f"The sum of selected orders is {total_sum}",
        level=messages.SUCCESS,
    )


@admin.action(description="Get the total items purchased")
def calculate_total_items(modeladmin, request, queryset):
    total_items = sum(order.total_items for order in queryset)
    modeladmin.message_user(
        request,
        f"The total number of items purchased is {total_items}",
        level=messages.SUCCESS,
    )


class OrderItemInline(admin.TabularInline):
    model = OrderItem
    form = OrderItemInlineForm
    formset = OrderItemInlineFormSet
    min_num = 1
    extra = 1
    readonly_fields = ["get_total_price"]
    fields = ["product", "quantity", "get_total_price"]
    autocomplete_fields = ["product"]

    def get_product_name(self, obj):
        """
        Display the stored historical item name (what was actually ordered),
        or current product name if no historical name is stored.
        """
        if obj:
            # Show stored historical name if available, otherwise current product name
            if obj.item_name:
                return obj.item_name
            elif obj.product:
                return obj.product.name
            else:
                return "Deleted product"
        return "-"
    
    get_product_name.short_description = "Item Name (Ordered)"

    def get_total_price(self, obj):
        """
        Calculate total price using stored price if product is deleted.
        """
        if obj:
            total = obj.get_total_price()
            if total != "":
                return f"£{total:.2f}"
        return "£0.00"

    get_total_price.short_description = "Total Price"


class ShippingDetailsInline(admin.StackedInline):
    model = ShippingDetails
    fk_name = "order"
    extra = 0
    can_delete = False
    readonly_fields = [
        "shipping_tracking_number",
        "shipping_tracking_url",
        "shipping_label_url",
        "sendcloud_parcel_id",
        "shipping_status",
        "shipping_error_message",
    ]
    fields = [
        "shipping_method_id",
        "shipping_carrier",
        "shipping_service_name",
        "shipping_cost",
        "shipping_tracking_number",
        "shipping_tracking_url",
        "shipping_label_url",
        "sendcloud_parcel_id",
        "shipping_status",
        "shipping_error_message",
    ]


def retry_shipment_creation(modeladmin, request, queryset):
    """
    Admin action to retry shipment creation for selected orders.
    """
    import logging

    from shipping.service import ShippingService

    logger = logging.getLogger(__name__)
    shipping_service = ShippingService()

    success_count = 0
    failed_count = 0
    skipped_count = 0

    for order in queryset:
        details = getattr(order, "shipping_details", None)
        # Skip orders that are not paid
        if order.status != "paid":
            skipped_count += 1
            continue

        # Skip orders without shipping method
        if not details or not details.shipping_method_id:
            skipped_count += 1
            continue

        try:
            result = shipping_service.create_shipment_for_order(order)
            if result.get("success"):
                success_count += 1
                logger.info(
                    f"Retry: Successfully created shipment for order {order.id}"
                )
            else:
                failed_count += 1
                logger.warning(
                    f"Retry: Failed to create shipment for order {order.id}: {result.get('error')}"
                )
        except Exception as e:
            failed_count += 1
            logger.error(
                f"Retry: Exception creating shipment for order {order.id}: {e}",
                exc_info=True,
            )

    # Show feedback message
    messages = []
    if success_count > 0:
        messages.append(f"{success_count} shipment(s) created successfully")
    if failed_count > 0:
        messages.append(f"{failed_count} shipment(s) failed")
    if skipped_count > 0:
        messages.append(
            f"{skipped_count} order(s) skipped (not paid or no shipping method)"
        )

    modeladmin.message_user(request, ". ".join(messages))


retry_shipment_creation.short_description = (
    "Retry shipment creation for selected orders"
)


@admin.register(Order)
class OrderAdmin(admin.ModelAdmin):
    form = OrderAdminForm

    class Media:
        js = (
            "admin/js/order_filter_cleaner.js",
            "admin/js/prevent_double_submit.js",
        )

    # Class-level font configuration cache (reused across all PDF exports)
    # This significantly reduces font processing time
    _font_config = None
    _font_cache_dir = None

    @classmethod
    def _get_font_config(cls):
        """Get or create shared FontConfiguration with persistent cache."""
        if cls._font_config is None:
            import logging
            from pathlib import Path

            from weasyprint.text.fonts import FontConfiguration

            # Setup persistent font cache directory
            cache_dir = Path("/tmp/weasyprint_cache")
            cache_dir.mkdir(parents=True, exist_ok=True)

            cls._font_cache_dir = cache_dir
            cls._font_config = FontConfiguration()

        return cls._font_config, cls._font_cache_dir

    actions = [
        create_and_upload_invoice,
        create_credit_note_and_void_invoice,
        mark_orders_paid,
        mark_orders_cancelled,
        mark_orders_pending,
        calculate_sum,
        calculate_total_items,
        # retry_shipment_creation,
        "export_orders_pdf",
        "food_summary_excel",
    ]

    list_display = [
        "id",
        "delivery_date",
        "customer_name",
        "customer_phone",
        "customer_address",
        "notes",
        "get_total_items",
        "get_total_price",
        "status",
        "get_invoice",
    ]
    list_display_links = ["id", "delivery_date", "customer_name"]
    list_filter = [
        DateFilter,
        "status",
        HolidayFeeFilter,
    ]
    list_editable = ["status"]
    search_fields = [
        "customer__name",
        "customer__profile__phone",
        "customer__email",
    ]
    ordering = ["-id"]
    date_hierarchy = "delivery_date"
    inlines = [OrderItemInline]
    autocomplete_fields = ["customer"]

    def get_inlines(self, request, obj):
        """Conditionally include ShippingDetailsInline only for orders with is_home_delivery=False."""
        inlines = [OrderItemInline]
        # Only show shipping details for orders that are NOT home delivery
        if obj and not obj.is_home_delivery:
            inlines.append(ShippingDetailsInline)
        return inlines

    def get_fields(self, request, obj=None):
        fields = [
            "customer",
            "notes",
            "status",
            "delivery_date",
            "delivery_date_order_id",
        ]
        if obj:
            fields += [
                "customer_phone",
                "customer_address",
                "get_total_items",
                "get_holiday_fee_amount",
                "get_total_price",
                "get_invoice",
                "is_home_delivery",
            ]
            if not request.user.has_perm("account.change_customuser"):
                return fields[0].replace("customer", "customer_name")
        fields += [
            "delivery_fee_manual",
            "delivery_fee",
            "holiday_fee",
            "discount",
        ]
        return fields

    def get_readonly_fields(self, request, obj=None):
        readonly = [
            "customer_name",
            "customer_phone",
            "customer_address",
            "get_total_items",
            "get_holiday_fee_amount",
            "get_total_price",
            "get_invoice",
            # Shipping fields are read-only (managed by system)
            "get_shipping_tracking_link",
            "get_shipping_label_link",
        ]
        # Make delivery_date_order_id readonly for non-superusers
        if not request.user.is_superuser:
            readonly.append("delivery_date_order_id")
        return readonly  # Admins can edit customer, status, notes

    def get_queryset(self, request):
        self.request = request  # Save request for later use
        qs = super().get_queryset(request)
        # Prefetch invoice to avoid N+1 queries when displaying invoice column
        return qs.select_related("invoice")

    def _is_single_date_filtered(self, request):
        """
        Check if a specific single date is filtered (not a date range).
        Returns True if:
        - DateFilter is set to "today"
        - date_hierarchy has all three parameters (year, month, day)
        """
        # Check DateFilter parameter
        date_filter_value = request.GET.get("future_date")
        if date_filter_value == "today":
            return True

        # Check date_hierarchy parameters (delivery_date__year, delivery_date__month, delivery_date__day)
        # If all three are present, it's a specific date
        year = request.GET.get("delivery_date__year")
        month = request.GET.get("delivery_date__month")
        day = request.GET.get("delivery_date__day")

        if year and month and day:
            return True

        return False

    def get_list_display(self, request):
        """
        Dynamically show delivery_date_order_id instead of id when a specific date is filtered.
        Show id for wider date ranges.
        """
        list_display = list(super().get_list_display(request))

        is_single_date = self._is_single_date_filtered(request)

        # If single date is filtered, replace 'id' with 'delivery_date_order_id'
        if is_single_date:
            if "id" in list_display:
                id_index = list_display.index("id")
                list_display[id_index] = "delivery_date_order_id"
        else:
            # For wider ranges, ensure 'id' is shown (it's already in the default list_display)
            if "delivery_date_order_id" in list_display:
                delivery_id_index = list_display.index("delivery_date_order_id")
                list_display[delivery_id_index] = "id"

        return list_display

    def get_list_display_links(self, request, list_display):
        """
        Dynamically update list_display_links based on whether we're showing id or delivery_date_order_id.
        """
        links = list(super().get_list_display_links(request, list_display))

        is_single_date = self._is_single_date_filtered(request)

        # Replace 'id' with 'delivery_date_order_id' in links if single date is filtered
        if is_single_date:
            links = [
                "delivery_date_order_id" if link == "id" else link for link in links
            ]

        return links

    def get_ordering(self, request):
        """
        Sort by delivery_date_order_id (descending) when a specific date is filtered.
        Otherwise use default ordering.
        """
        is_single_date = self._is_single_date_filtered(request)

        if is_single_date:
            # Sort by delivery_date_order_id descending (last to first)
            return ["-delivery_date_order_id"]

        # Default ordering
        return super().get_ordering(request) or ["-id"]

    def customer_name(self, obj):
        request = getattr(self, "request", None)
        if request and request.user.has_perm("account.change_customuser"):
            if obj.customer:
                url = reverse("admin:account_customuser_change", args=[obj.customer.id])
                return format_html('<a href="{}">{}</a>', url, obj.customer.name)
        return obj.customer.name if obj.customer else "No Customer"

    def customer_phone(self, obj):
        return (
            obj.customer.profile.phone
            if obj.customer and obj.customer.profile
            else "No Phone"
        )

    def save_related(self, request, form, formsets, change):
        super().save_related(request, form, formsets, change)
        order = form.instance
        items = order.items.all()

        if not order.delivery_fee_manual:
            from shipping.service import ShippingService

            # Sausage category name
            post_suitable_category = "Sausages and Marinated products"
            for item in items:
                category_names = item.product.categories.values_list("name", flat=True)
                if post_suitable_category not in [
                    name.lower() for name in category_names
                ]:
                    order.is_home_delivery = True
                    order.delivery_fee = 10
                    break
            else:
                order.is_home_delivery = False
                if order.total_price > 220:
                    order.delivery_fee = 0
                else:
                    # Use Royal Mail pricing (same as cart)
                    total_weight = sum(item.quantity for item in items)
                    order.delivery_fee = ShippingService.get_delivery_fee_by_weight(
                        float(total_weight)
                    )
        order.save()

    def get_total_price(self, obj):
        return obj.total_price

    get_total_price.short_description = "Total Price"

    def get_total_items(self, obj):
        return obj.total_items

    get_total_items.short_description = "Total Items"

    def get_holiday_fee_amount(self, obj):
        """Display holiday fee in actual money value, only if holiday_fee > 0."""
        if obj.holiday_fee > 0:
            return f"{obj.holiday_fee_amount:.2f}"
        return "-"

    get_holiday_fee_amount.short_description = "Holiday Fee (£)"

    def get_shipping_status_display(self, obj):
        """Display shipping status with color coding."""
        details = getattr(obj, "shipping_details", None)
        if not details or not details.shipping_status:
            if obj.status == "paid" and details and details.shipping_method_id:
                return format_html('<span style="color: orange;">⏳ Pending</span>')
            return "-"

        status_colors = {
            "pending_shipment": "orange",
            "label_created": "green",
            "shipment_failed": "red",
            "in_transit": "blue",
            "out_for_delivery": "purple",
            "delivered": "darkgreen",
        }

        status_icons = {
            "pending_shipment": "⏳",
            "label_created": "✅",
            "shipment_failed": "❌",
            "in_transit": "🚚",
            "out_for_delivery": "📦",
            "delivered": "✓",
        }

        color = status_colors.get(details.shipping_status, "gray")
        icon = status_icons.get(details.shipping_status, "")
        label = details.get_shipping_status_display()

        return format_html('<span style="color: {};">{} {}</span>', color, icon, label)

    get_shipping_status_display.short_description = "Shipping Status"

    def get_shipping_tracking_link(self, obj):
        """Display tracking number as clickable link if available."""
        details = getattr(obj, "shipping_details", None)
        if details and details.shipping_tracking_number:
            if details.shipping_tracking_url:
                return format_html(
                    '<a href="{}" target="_blank">{}</a>',
                    details.shipping_tracking_url,
                    details.shipping_tracking_number,
                )
            return details.shipping_tracking_number
        return "-"

    get_shipping_tracking_link.short_description = "Tracking Number"

    def get_shipping_label_link(self, obj):
        """Display label download link if available."""
        details = getattr(obj, "shipping_details", None)
        if details and details.shipping_label_url:
            return format_html(
                '<a href="{}" target="_blank" class="button">📄 Download Label</a>',
                details.shipping_label_url,
            )
        return "-"

    get_shipping_label_link.short_description = "Shipping Label"

    def save_model(self, request, obj, form, change):
        """
        Save the order without calculating delivery fees here to avoid double save.

        Note: The Order.save() method will automatically handle:
        - Auto-assignment of delivery_date_order_id for new orders
        - Reassignment of delivery_date_order_id when delivery_date changes
        This happens automatically when super().save_model() calls obj.save()
        """
        # Don't calculate delivery fees in save_model to prevent double save
        # This will be handled in save_related after inlines are processed
        # Delivery date order ID reassignment is handled in Order.save() method
        
        # Track if status changed to "paid" to update invoice
        status_changed_to_paid = False
        if change and "status" in form.changed_data and obj.status == "paid":
            status_changed_to_paid = True
        
        super().save_model(request, obj, form, change)
        
        # Update related invoice when order status changes to "paid"
        if status_changed_to_paid:
            self._update_invoice_to_paid(obj)

    def _update_invoice_to_paid(self, order):
        """
        Update the related invoice when order status changes to "paid".
        Sets amount_paid to total_amount, status to PAID, and paid_at to now.
        """
        from billing.models import Invoice
        from django.utils import timezone

        invoice = getattr(order, "invoice", None)
        if invoice and invoice.status != Invoice.Status.PAID:
            invoice.amount_paid = invoice.total_amount
            invoice.status = Invoice.Status.PAID
            invoice.paid_at = timezone.now()
            invoice.save(update_fields=["amount_paid", "status", "paid_at"])

    def save_formset(self, request, form, formset, change):
        """
        Save order items. Validation for duplicates is handled by OrderItemInlineFormSet.
        """
        if formset.model is OrderItem:
            # Validation happens in formset.clean(), so if we get here, it's valid
            instances = formset.save(commit=False)

            with transaction.atomic():
                for instance in instances:
                    if not instance.product or not instance.quantity:
                        continue
                    instance.order = form.instance
                    instance.save()

                # Handle deleted items
                for obj in formset.deleted_objects:
                    obj.delete()
        else:
            # For other formsets, use default behavior
            super().save_formset(request, form, formset, change)

    def save_related(self, request, form, formsets, change):
        """Calculate delivery fields after saving inlines and save only once."""
        super().save_related(request, form, formsets, change)
        order = form.instance

        # Only calculate delivery fees if not manually set
        if not order.delivery_fee_manual:
            # Build compute source from current items after inlines are saved
            compute_source = [
                {"product": i.product, "quantity": i.quantity}
                for i in order.items.all()
            ]

            # Calculate delivery fee and home status
            is_home_delivery, delivery_fee = (
                order.calculate_delivery_fee_and_home_status_from_items(compute_source)
            )

            # Only save if values have changed to prevent unnecessary saves
            if (
                order.is_home_delivery != is_home_delivery
                or order.delivery_fee != delivery_fee
            ):
                order.is_home_delivery = is_home_delivery
                order.delivery_fee = delivery_fee
                order.save(update_fields=["is_home_delivery", "delivery_fee"])

    def formfield_for_foreignkey(self, db_field, request, **kwargs):
        if db_field.name == "customer":
            kwargs["queryset"] = CustomUser.objects.filter(is_staff=False)
            if not request.user.has_perm("account.change_customuser"):
                kwargs["disabled"] = True
        return super().formfield_for_foreignkey(db_field, request, **kwargs)

    def get_invoice(self, obj):
        """
        Display invoice download link if exists, otherwise show "Create invoice" button.
        """
        invoice = getattr(obj, "invoice", None)
        if invoice and getattr(invoice, "invoice_link", ""):
            try:
                url = invoice.get_presigned_invoice_url(expires_in=300)
                return format_html(
                    '<a href="{}" target="_blank" class="button" style="background-color: #28a745; color: white; padding: 5px 10px; text-decoration: none; border-radius: 3px;">Download</a>',
                    url,
                )
            except Exception:
                # fallback below
                pass

        # Check for backward compat (old Order.invoice_link)
        if obj.invoice_link:
            s3 = get_s3_client()
            try:
                url = s3.generate_presigned_url(
                    "get_object",
                    Params={
                        "Bucket": settings.AWS_STORAGE_BUCKET_NAME,
                        "Key": obj.invoice_link,
                    },
                    ExpiresIn=300,
                )
                return format_html(
                    '<a href="{}" target="_blank" class="button" style="background-color: #28a745; color: white; padding: 5px 10px; text-decoration: none; border-radius: 3px;">Download</a>',
                    url,
                )
            except Exception:
                pass

        # No invoice exists - show "Create invoice" button
        create_url = reverse(
            "admin:api_order_create_invoice", kwargs={"order_id": obj.id}
        )
        return format_html(
            '<a href="{}" class="button" style="background-color: #417690; color: white; padding: 5px 10px; text-decoration: none; border-radius: 3px;">Create</a>',
            create_url,
        )

    get_invoice.short_description = "Invoice"

    def get_urls(self):
        """Add custom URL for single-order invoice creation."""
        urls = super().get_urls()
        custom_urls = [
            path(
                "<int:order_id>/create-invoice/",
                self.admin_site.admin_view(self.create_invoice_view),
                name="api_order_create_invoice",
            ),
        ]
        return custom_urls + urls

    def create_invoice_view(self, request, order_id):
        """
        Create invoice for a single order and redirect back to order list.
        """
        from django.db import transaction

        try:
            order = Order.objects.get(pk=order_id)
        except Order.DoesNotExist:
            self.message_user(
                request, f"Order {order_id} not found.", level=messages.ERROR
            )
            return HttpResponseRedirect(reverse("admin:api_order_changelist"))

        try:
            with transaction.atomic():
                from billing.models import Invoice

                invoice = Invoice.create_and_publish_from_order(
                    order=order, request=request
                )

                # Backward compat: keep the Order.invoice_link in sync
                if invoice.invoice_link and order.invoice_link != invoice.invoice_link:
                    order.invoice_link = invoice.invoice_link
                    order.save(update_fields=["invoice_link"])

                self.message_user(
                    request,
                    f"Invoice created successfully for Order #{order.id}.",
                    level=messages.SUCCESS,
                )
        except Exception as e:
            import logging

            logger = logging.getLogger(__name__)
            logger.error(
                f"Failed to create invoice for order {order_id}: {str(e)}",
                exc_info=True,
            )
            self.message_user(
                request,
                f"Failed to create invoice: {str(e)}",
                level=messages.ERROR,
            )

        # Redirect back to order list or order detail page
        return HttpResponseRedirect(reverse("admin:api_order_changelist"))

    def export_orders_pdf(self, request, queryset):
        """
        Export selected orders to PDF with optimized memory usage and performance.

        Optimizations:
        - Efficient database queries with select_related/prefetch_related
        - Font caching to avoid repeated font subsetting
        - Chunked processing for large datasets (reduces memory usage)
        - Streaming response for very large exports
        - Memory-efficient PDF generation settings
        """
        import logging
        import time

        start_time = time.time()

        # Validate queryset
        order_count = queryset.count()
        if order_count == 0:
            self.message_user(request, "No orders selected.", level=messages.WARNING)
            return None

        # Estimate processing time (roughly 1-2 seconds per order for PDF export)
        estimated_time = order_count * 0.8
        logger = logging.getLogger(__name__)
        logger.info(
            f"Starting PDF creation for {order_count} order(s), "
            f"approximate time {estimated_time:.1f} seconds"
        )

        try:
            # Check for PDF merger availability
            try:
                from pypdf import PdfReader, PdfWriter

                pdf_merger_available = True
            except ImportError:
                pdf_merger_available = False
                if order_count > 50:
                    messages.warning(
                        request,
                        "Warning: pypdf not installed. Large exports may be slow. "
                        "Install pypdf for better performance: pip install pypdf",
                    )

            # Optimize queryset to avoid N+1 queries
            from django.db.models import Prefetch

            optimized_queryset = (
                queryset.select_related("customer", "customer__profile")
                .prefetch_related(
                    Prefetch(
                        "items",
                        queryset=OrderItem.objects.select_related("product"),
                        to_attr="prefetched_items",
                    )
                )
                .order_by("delivery_date", "id")
            )

            # Get shared font configuration
            font_config, cache_dir = self._get_font_config()

            # Generate filename
            filename = self._generate_pdf_filename(queryset, order_count)

            # Generate PDF
            response = self._generate_pdf_by_strategy(
                optimized_queryset=optimized_queryset,
                order_count=order_count,
                font_config=font_config,
                cache_dir=cache_dir,
                filename=filename,
                request=request,
                pdf_merger_available=pdf_merger_available,
            )

            # Show success message - stored in session, will show when user returns to admin page
            # elapsed_time = time.time() - start_time
            # self._show_pdf_success_message(request, order_count, elapsed_time)

            return response

        except Exception as e:
            # Log error and show user-friendly message
            import logging

            logger = logging.getLogger(__name__)
            elapsed_time = time.time() - start_time
            error_message = (
                f"Failed to create PDF after {elapsed_time:.1f} seconds: {str(e)}"
            )
            logger.error(error_message, exc_info=True)
            self.message_user(
                request,
                error_message,
                level=messages.ERROR,
            )
            return None

    def _generate_pdf_filename(self, queryset, order_count):
        """Generate appropriate filename for PDF export."""
        delivery_dates = queryset.values_list("delivery_date", flat=True).distinct()
        if len(delivery_dates) == 1:
            return f"{delivery_dates[0].strftime('%d-%b-%Y')}_Orders.pdf"
        return f"Orders_{order_count}_items.pdf"

    def _get_chunk_size(self, order_count):
        """Calculate optimal chunk size based on order count."""
        if order_count <= 50:
            return 50
        elif order_count <= 200:
            return 100
        elif order_count <= 500:
            return 150
        else:
            return 200

    def _generate_pdf_by_strategy(
        self,
        optimized_queryset,
        order_count,
        font_config,
        cache_dir,
        filename,
        request,
        pdf_merger_available,
    ):
        """
        Select and execute the appropriate PDF generation strategy.

        Returns:
            HttpResponse or StreamingHttpResponse
        """
        STREAMING_THRESHOLD = 100
        chunk_size = self._get_chunk_size(order_count)

        # Strategy 1: Single PDF (no merger available or small dataset)
        if not pdf_merger_available or order_count <= chunk_size:
            return self._generate_single_pdf(
                optimized_queryset, font_config, cache_dir, filename, request
            )

        # Strategy 2: Chunked and merged (medium datasets)
        if order_count <= STREAMING_THRESHOLD:
            return self._generate_chunked_pdf(
                optimized_queryset,
                order_count,
                chunk_size,
                font_config,
                cache_dir,
                filename,
                request,
            )

        # Strategy 3: Streaming (large datasets)
        return self._generate_streaming_pdf(
            optimized_queryset,
            order_count,
            chunk_size,
            font_config,
            cache_dir,
            filename,
            request,
        )

    def _show_pdf_success_message(self, request, order_count, elapsed_time):
        """Display success message after PDF generation."""
        import logging

        logger = logging.getLogger(__name__)
        message = f"Successfully created PDF with {order_count} order(s) in {elapsed_time:.1f} seconds."
        logger.info(message)
        # Store message in session - message_user handles this, but ensure it's called
        self.message_user(
            request,
            message,
            level=messages.SUCCESS,
        )

    def _generate_single_pdf(self, queryset, font_config, cache_dir, filename, request):
        """Generate PDF for small exports - fastest method."""
        import gc

        from weasyprint import HTML

        # Convert to list to evaluate prefetch
        orders_list = list(queryset)

        # Pre-process data
        for order in orders_list:
            if hasattr(order, "prefetched_items"):
                order._items_cache = order.prefetched_items
            order._cached_total_price = order.total_price

        html_string = render_to_string("orders.html", {"orders": orders_list})

        with tempfile.NamedTemporaryFile(delete=True, suffix=".pdf") as output:
            HTML(
                string=html_string,
                base_url=request.build_absolute_uri("/"),
            ).write_pdf(
                target=output.name,
                font_config=font_config,
                optimize_images=True,
                jpeg_quality=85,
                dpi=150,
                cache=cache_dir,
            )

            gc.collect()
            output.seek(0)
            pdf_content = output.read()

            response = HttpResponse(pdf_content, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            del pdf_content, html_string, orders_list
            gc.collect()

            return response

    def _generate_chunked_pdf(
        self,
        queryset,
        order_count,
        chunk_size,
        font_config,
        cache_dir,
        filename,
        request,
    ):
        """Generate PDF in chunks and merge - memory efficient for medium exports."""
        import gc
        import os
        import shutil

        from pypdf import PdfReader, PdfWriter
        from weasyprint import HTML

        # Create temporary directory for chunk PDFs
        temp_dir = tempfile.mkdtemp()
        chunk_files = []

        try:
            # Optimize PDF settings for very large exports
            # Reduce DPI and quality for faster processing on large exports
            use_fast_mode = order_count > 500
            pdf_dpi = 120 if use_fast_mode else 150
            jpeg_quality = 75 if use_fast_mode else 85

            # Process orders in chunks
            total_chunks = (order_count + chunk_size - 1) // chunk_size
            for chunk_idx, i in enumerate(range(0, order_count, chunk_size), 1):
                chunk_orders = queryset[i : i + chunk_size]
                orders_list = list(chunk_orders)

                # Pre-process chunk data
                for order in orders_list:
                    if hasattr(order, "prefetched_items"):
                        order._items_cache = order.prefetched_items
                    order._cached_total_price = order.total_price

                # Generate HTML for chunk
                html_string = render_to_string("orders.html", {"orders": orders_list})

                # Generate PDF chunk with optimized settings
                chunk_file = os.path.join(temp_dir, f"chunk_{i}.pdf")
                HTML(
                    string=html_string,
                    base_url=request.build_absolute_uri("/"),
                ).write_pdf(
                    target=chunk_file,
                    font_config=font_config,
                    optimize_images=True,
                    jpeg_quality=jpeg_quality,
                    dpi=pdf_dpi,
                    cache=cache_dir,
                )

                chunk_files.append(chunk_file)

                # Clean up chunk data immediately
                del html_string, orders_list
                gc.collect()

            # Merge all PDF chunks using PdfWriter (pypdf 5.0+)
            # Optimize merging: read and append in one pass
            writer = PdfWriter()
            for chunk_file in chunk_files:
                with open(chunk_file, "rb") as f:
                    reader = PdfReader(f)
                    writer.append(reader)
                    # Close file immediately to free memory
                # Delete chunk file immediately after reading to save disk space
                if os.path.exists(chunk_file):
                    os.unlink(chunk_file)

            # Write merged PDF to temporary file
            with tempfile.NamedTemporaryFile(
                delete=False, suffix=".pdf", mode="wb"
            ) as merged_output:
                writer.write(merged_output)
                merged_output.flush()  # Ensure data is written

                # Read merged PDF
                with open(merged_output.name, "rb") as f:
                    pdf_content = f.read()

                # Clean up (chunk files already deleted during merge)
                os.unlink(merged_output.name)

            response = HttpResponse(pdf_content, content_type="application/pdf")
            response["Content-Disposition"] = f'attachment; filename="{filename}"'

            del pdf_content
            gc.collect()

            return response

        finally:
            # Cleanup temp directory
            if os.path.exists(temp_dir):
                shutil.rmtree(temp_dir, ignore_errors=True)

    def _generate_streaming_pdf(
        self,
        queryset,
        order_count,
        chunk_size,
        font_config,
        cache_dir,
        filename,
        request,
    ):
        """Generate and stream PDF chunks - most memory efficient for large exports."""
        import gc
        import os
        import shutil

        from django.http import StreamingHttpResponse
        from pypdf import PdfReader, PdfWriter
        from weasyprint import HTML

        def pdf_generator():
            """Generator that yields PDF chunks."""
            # Optimize PDF settings for very large exports
            use_fast_mode = order_count > 500
            pdf_dpi = 120 if use_fast_mode else 150
            jpeg_quality = 75 if use_fast_mode else 85

            temp_dir = tempfile.mkdtemp()
            chunk_files = []

            try:
                # Process first chunk
                first_chunk = queryset[0:chunk_size]
                orders_list = list(first_chunk)

                for order in orders_list:
                    if hasattr(order, "prefetched_items"):
                        order._items_cache = order.prefetched_items
                    order._cached_total_price = order.total_price

                html_string = render_to_string("orders.html", {"orders": orders_list})
                chunk_file = os.path.join(temp_dir, f"chunk_0.pdf")

                HTML(
                    string=html_string,
                    base_url=request.build_absolute_uri("/"),
                ).write_pdf(
                    target=chunk_file,
                    font_config=font_config,
                    optimize_images=True,
                    jpeg_quality=jpeg_quality,
                    dpi=pdf_dpi,
                    cache=cache_dir,
                )

                chunk_files.append(chunk_file)
                del html_string, orders_list
                gc.collect()

                # Process remaining chunks
                for i in range(chunk_size, order_count, chunk_size):
                    chunk_orders = queryset[i : i + chunk_size]
                    orders_list = list(chunk_orders)

                    for order in orders_list:
                        if hasattr(order, "prefetched_items"):
                            order._items_cache = order.prefetched_items
                        order._cached_total_price = order.total_price

                    html_string = render_to_string(
                        "orders.html", {"orders": orders_list}
                    )
                    chunk_file = os.path.join(temp_dir, f"chunk_{i}.pdf")

                    HTML(
                        string=html_string,
                        base_url=request.build_absolute_uri("/"),
                    ).write_pdf(
                        target=chunk_file,
                        font_config=font_config,
                        optimize_images=True,
                        jpeg_quality=jpeg_quality,
                        dpi=pdf_dpi,
                        cache=cache_dir,
                    )

                    chunk_files.append(chunk_file)
                    del html_string, orders_list
                    gc.collect()

                # Merge and stream using PdfWriter (pypdf 5.0+)
                # Optimize merging: read and append in one pass, delete immediately
                writer = PdfWriter()
                for chunk_file in chunk_files:
                    with open(chunk_file, "rb") as f:
                        reader = PdfReader(f)
                        writer.append(reader)
                    # Delete chunk file immediately after reading to save disk space
                    if os.path.exists(chunk_file):
                        os.unlink(chunk_file)

                with tempfile.NamedTemporaryFile(
                    delete=False, suffix=".pdf", mode="wb"
                ) as merged_output:
                    writer.write(merged_output)
                    merged_output.flush()  # Ensure data is written

                    # Stream the merged PDF in chunks
                    with open(merged_output.name, "rb") as f:
                        while True:
                            chunk = f.read(8192)  # 8KB chunks
                            if not chunk:
                                break
                            yield chunk

                    os.unlink(merged_output.name)

            finally:
                # Cleanup (chunk files already deleted during merge)
                if os.path.exists(temp_dir):
                    shutil.rmtree(temp_dir, ignore_errors=True)

        response = StreamingHttpResponse(
            pdf_generator(), content_type="application/pdf"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        response["Content-Type"] = "application/pdf"

        return response

    export_orders_pdf.short_description = "Export Selected Orders to PDF"

    def food_summary_csv(self, request, queryset):
        """
        Export a food summary of all products and their total quantities for selected orders as CSV.
        Frozen products are on the left, fresh products on the right.
        """
        order_items = OrderItem.objects.filter(order__in=queryset).select_related(
            "product"
        )
        frozen_summary = {}
        ready_summary = {}
        for item in order_items:
            product = item.product
            if not product:
                continue
            if product.categories.filter(name__iexact="Frozen Products").exists():
                frozen_summary[product.name] = frozen_summary.get(
                    product.name, 0
                ) + float(item.quantity)
            else:
                ready_summary[product.name] = ready_summary.get(
                    product.name, 0
                ) + float(item.quantity)

        # Sort and convert to lists of tuples
        frozen_list = sorted(frozen_summary.items())
        ready_list = sorted(ready_summary.items())

        # Pad the shorter list
        max_len = max(len(frozen_list), len(ready_list))
        frozen_list += [("", "")] * (max_len - len(frozen_list))
        ready_list += [("", "")] * (max_len - len(ready_list))

        delivery_dates = queryset.values_list("delivery_date", flat=True).distinct()
        response = HttpResponse(content_type="text/csv")
        response["Content-Disposition"] = (
            f'attachment; filename="{", ".join(sorted({d.strftime("%d-%b-%Y") for d in delivery_dates}))}_Products.csv"'
        )
        writer = csv.writer(response)
        writer.writerow(["Frozen Product", "Quantity", "", "Ready Product", "Quantity"])
        for (f_name, f_qty), (r_name, r_qty) in zip(frozen_list, ready_list):
            writer.writerow([f_name, f_qty, "", r_name, r_qty])
        return response

    food_summary_csv.short_description = "Export Food Summary as CSV"

    def food_summary_excel(self, request, queryset):
        """
        Export a food summary of all products and their total quantities for selected orders as Excel.
        Frozen products are on the left, fresh products on the right.
        """
        order_items = OrderItem.objects.filter(order__in=queryset).select_related(
            "product"
        )

        frozen_summary = {}
        ready_summary = {}

        for item in order_items:
            product = item.product
            if not product:
                continue
            if product.categories.filter(name__iexact="Frozen Products").exists():
                frozen_summary[product.name] = frozen_summary.get(
                    product.name, 0
                ) + float(item.quantity)
            else:
                ready_summary[product.name] = ready_summary.get(
                    product.name, 0
                ) + float(item.quantity)

        frozen_list = sorted(frozen_summary.items())
        ready_list = sorted(ready_summary.items())

        max_len = max(len(frozen_list), len(ready_list))
        frozen_list += [("", "")] * (max_len - len(frozen_list))
        ready_list += [("", "")] * (max_len - len(ready_list))

        delivery_dates = queryset.values_list("delivery_date", flat=True).distinct()
        filename = (
            ", ".join(sorted({d.strftime("%d-%b-%Y") for d in delivery_dates}))
            + "_Products.xlsx"
        )

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Food Summary"

        ws.append(["Frozen Product", "Quantity", "", "Ready Product", "Quantity"])

        for (f_name, f_qty), (r_name, r_qty) in zip(frozen_list, ready_list):
            ws.append([f_name, f_qty, "", r_name, r_qty])

        # Auto-size columns
        for col in ws.columns:
            max_length = max(len(str(cell.value or "")) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max_length + 3

        # Return HTTP response
        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        wb.save(response)
        return response

    food_summary_excel.short_description = "Export Food Summary as Excel"
